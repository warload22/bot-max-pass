import io
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from app.core.config import Config
from app.repositories.registration import RegistrationRepository
from app.repositories.anonymized_stat import AnonymizedStatRepository
from app.handlers.common import get_category_display, get_education_display

def export_full_registrations(db, event_id: int) -> bytes:
    """Генерирует Excel с полными данными регистраций (время регистрации с учётом TIME_OFFSET)."""
    reg_repo = RegistrationRepository(db)
    registrations = reg_repo.get_all_by_event(event_id)

    if not registrations:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Регистрации"

    headers = [
        "UUID", "Дата регистрации", "ФИО", "Дата рождения", "Место рождения",
        "Место жительства", "Email", "Категория", "Интерес", "Учебное заведение",
        "Активна", "Последняя отправка QR", "Статус сканирования"
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')

    from app.repositories.scan import ScanRepository
    scan_repo = ScanRepository(db)

    for reg in registrations:
        last_scan = scan_repo.get_last_by_registration(reg.id)
        if last_scan:
            if last_scan.status == 'admitted':
                status_display = 'Пропущен'
            elif last_scan.status == 'denied':
                status_display = 'Не пропущен'
            else:
                status_display = 'Не сканирован'
        else:
            status_display = 'Не сканирован'

        # Применяем смещение времени для отображения в локальном часовом поясе
        if reg.registered_at:
            local_time = reg.registered_at + timedelta(seconds=Config.TIME_OFFSET)
            registered_at_str = local_time.strftime('%d.%m.%Y %H:%M')
        else:
            registered_at_str = ''

        if reg.last_qr_sent_at:
            last_qr_local = reg.last_qr_sent_at + timedelta(seconds=Config.TIME_OFFSET)
            last_qr_str = last_qr_local.strftime('%d.%m.%Y %H:%M')
        else:
            last_qr_str = ''

        row = [
            str(reg.id),
            registered_at_str,
            reg.full_name,
            reg.birth_date.strftime('%d.%m.%Y') if reg.birth_date else '',
            reg.birth_place,
            reg.residence or '',
            reg.email or '',
            get_category_display(reg.category) if reg.category else '',
            get_education_display(reg.education_interest) if reg.education_interest else '',
            reg.school or '',
            'Да' if reg.is_active else 'Нет',
            last_qr_str,
            status_display
        ]
        ws.append(row)

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_anonymized_stats(db, event_id: int) -> bytes:
    """Генерирует Excel с обезличенной статистикой (время регистрации с учётом TIME_OFFSET)."""
    anon_repo = AnonymizedStatRepository(db)
    stats = anon_repo.get_by_event(event_id)

    if not stats:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Обезличенная статистика"

    headers = [
        "Год рождения", "Место рождения", "Место жительства",
        "Категория", "Интерес", "Учебное заведение",
        "Статус сканирования", "Дата регистрации"
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')

    for stat in stats:
        if stat.scan_status == 'admitted':
            status_display = 'Пропущен'
        elif stat.scan_status == 'denied':
            status_display = 'Не пропущен'
        else:
            status_display = 'Не сканирован'

        if stat.registered_at:
            local_time = stat.registered_at + timedelta(seconds=Config.TIME_OFFSET)
            reg_at_str = local_time.strftime('%d.%m.%Y %H:%M')
        else:
            reg_at_str = ''

        row = [
            stat.birth_year or '',
            stat.birth_place or '',
            stat.residence or '',
            get_category_display(stat.category) if stat.category else '',
            get_education_display(stat.education_interest) if stat.education_interest else '',
            stat.school or '',
            status_display,
            reg_at_str
        ]
        ws.append(row)

    # Автоширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()